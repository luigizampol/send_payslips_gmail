import smtplib
import ssl
from email.message import EmailMessage
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import os.path
import openpyxl

def extract_emails_from_xlsx_file(file_name, name_column_name, email_column_name):
    
    # Create a dictionary to store name-email pairs
    name_email_pairs = {}
    
    # Open the Excel file
    try:
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
    except Exception as e:
        print(f"Error opening the file: {e}")
        return name_email_pairs
    
    # Map the indices of name and email columns
    name_column_index = None
    email_column_index = None
    for col_idx, col in enumerate(sheet.iter_cols(min_row=1, max_row=1, values_only=True), start=1):
        if col[0] == name_column_name:
            name_column_index = col_idx
        elif col[0] == email_column_name:
            email_column_index = col_idx
    
    # Check if the columns were found
    if name_column_index is None or email_column_index is None:
        print(f"Columns '{name_column_name}' and '{email_column_name}' were not found.")
        wb.close()
        return name_email_pairs
    
    # Iterate through the rows of the file
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[name_column_index - 1]
        email = row[email_column_index - 1]
        
        # Check if the email value is present
        if email:
            name_email_pairs[name] = email
    
    # Close the file
    wb.close()
    
    return name_email_pairs

# Name of the Excel file containing employee names and emails (should be in the same directory as the script)
file_name = "employees.xlsx"

# Define column names
name_column = "name"  # Change this to the actual name of the name column
email_column = "email"  # Change this to the actual name of the email column

# Call the function to extract name-email pairs
pairs = extract_emails_from_xlsx_file(file_name, name_column, email_column)

# Define email settings
email_login = 'company@gmail.com'
email_sender = 'company@gmail.com'
email_password = 'osaidjpmismd'

# Define the email subject and body
payslip_description = 'August Salary'
subject = payslip_description
body = f"""
<html>
<head></head>
<body>
    <p>Dear Employee,</p>
    <p>We are sending your payslip for {payslip_description}</p>
    <p>You can view the payslip in the attached PDF file to this email.</p>
    <p>If you have any questions or need more information, feel free to contact us.</p>
    <p>We appreciate your work and dedication.</p>
    <p>Best regards,</p>
    <p>Company Name</p>
</body>
</html>
"""

# Define folder names
assets_folder = "Active employees"  # Change this to your desired assets folder name
payslips_folder = "Payslips"  # Change this to your desired payslips folder name. This folder should be inside each employee's folder. Each employee's folder must be named after the employee. The employee folders should be within the folder mentioned in the variable 'assets_folder'.

# Create the SSL context
context = ssl.create_default_context()

# Connect to the Gmail SMTP server
with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
    smtp.login(email_login, email_password)

    for employee_name, employee_email in pairs.items():
        employee_folder_path = os.path.join(assets_folder, employee_name)
        payslip_folder_path = os.path.join(employee_folder_path, payslips_folder)

        if os.path.exists(payslip_folder_path):
            # Find the latest PDF file in the payslip folder
            latest_pdf_file = max(
                os.listdir(payslip_folder_path),
                key=lambda f: os.path.getmtime(os.path.join(payslip_folder_path, f))
            )
            pdf_file_path = os.path.join(payslip_folder_path, latest_pdf_file)

            # Create a multipart message
            msg = MIMEMultipart()
            msg['From'] = email_sender
            msg['To'] = employee_email
            msg['Subject'] = subject

            # Attach the HTML body
            html_part = MIMEText(body, 'html')
            msg.attach(html_part)

            # Attach the PDF file to the email
            with open(pdf_file_path, 'rb') as attachment:
                pdf_attachment = MIMEApplication(attachment.read(), _subtype='pdf')
                pdf_attachment.add_header('content-disposition', f'attachment', filename=os.path.basename(pdf_file_path))
                msg.attach(pdf_attachment)

            # Send the email to the employee
            smtp.sendmail(email_sender, employee_email, msg.as_string())
            print(f"Email sent to {employee_name} ({employee_email}) successfully!")

print("All emails have been sent.")
